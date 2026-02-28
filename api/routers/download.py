"""Download API -- Advertiser report Excel, creative images ZIP, advertiser list CSV, gallery selection ZIP."""

import csv
import io
import logging
import os
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from api.deps import get_current_user, require_paid, require_plan
from database import get_db
from database.models import (
    AdDetail,
    AdSnapshot,
    Advertiser,
    Campaign,
    Industry,
    SpendEstimate,
    User,
)

router = APIRouter(
    prefix="/api/download",
    tags=["download"],
    dependencies=[Depends(require_paid)],
)

logger = logging.getLogger("adscope.download")

KST = timezone(timedelta(hours=9))
IMAGE_STORE_DIR = Path(os.getenv("IMAGE_STORE_DIR", "stored_images"))
SCREENSHOTS_DIR = Path("screenshots")

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

UTF8_BOM = "\ufeff"

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _safe(value) -> str:
    if value is None:
        return ""
    return str(value)


def _kst_str(dt: datetime | None) -> str:
    if dt is None:
        return ""
    kst_dt = dt + timedelta(hours=9)
    return kst_dt.strftime("%Y-%m-%d %H:%M:%S")


def _kst_date_str(dt: datetime | None) -> str:
    if dt is None:
        return ""
    kst_dt = dt + timedelta(hours=9)
    return kst_dt.strftime("%Y-%m-%d")


def _today_str() -> str:
    return datetime.now(KST).strftime("%Y%m%d")


def _safe_content_disposition(filename: str) -> dict[str, str]:
    """Build Content-Disposition header that supports non-ASCII filenames (RFC 5987)."""
    # ASCII fallback: replace non-ASCII chars with underscores
    ascii_name = filename.encode("ascii", "replace").decode("ascii").replace("?", "_")
    # UTF-8 encoded filename
    utf8_name = quote(filename, safe="")
    return {
        "Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"
    }


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws, col_count: int, max_width: int = 40):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 10)


def _default_date_range(date_from: datetime | None, date_to: datetime | None):
    if date_to is None:
        date_to = datetime.utcnow()
    if date_from is None:
        date_from = date_to - timedelta(days=30)
    return date_from, date_to


def _resolve_image_path(rel_path: str | None) -> Path | None:
    """Resolve a relative creative_image_path to an absolute filesystem path.

    Images may be stored in stored_images/ or screenshots/ directories.
    The rel_path in the DB may contain the prefix directory or not.

    Security: blocks path traversal attacks (.. sequences) and ensures
    the resolved path stays within allowed sandbox directories.
    """
    if not rel_path:
        return None

    # Normalise separators and strip leading slashes
    clean = rel_path.replace("\\", "/").lstrip("/")

    # Block any path-traversal attempt
    if ".." in clean:
        return None

    # Allowed sandbox roots (resolved to absolute)
    root = Path(os.getcwd()).resolve()
    allowed_roots = [
        root,
        IMAGE_STORE_DIR.resolve(),
        SCREENSHOTS_DIR.resolve(),
    ]

    def _is_within_sandbox(p: Path) -> bool:
        """Return True if *p* lives under one of the allowed roots."""
        resolved = p.resolve()
        for allowed in allowed_roots:
            try:
                resolved.relative_to(allowed)
                return True
            except ValueError:
                continue
        return False

    # Try relative to project root
    full = (root / clean).resolve()
    if _is_within_sandbox(full) and full.exists():
        return full

    # Try under stored_images/
    full = (IMAGE_STORE_DIR / clean).resolve()
    if _is_within_sandbox(full) and full.exists():
        return full

    # Try under screenshots/
    full = (SCREENSHOTS_DIR / clean).resolve()
    if _is_within_sandbox(full) and full.exists():
        return full

    # If the path starts with "stored_images/" or "screenshots/" strip it and retry
    for prefix_dir in (IMAGE_STORE_DIR, SCREENSHOTS_DIR):
        prefix_str = str(prefix_dir).replace("\\", "/") + "/"
        if clean.startswith(prefix_str):
            inner = clean[len(prefix_str):]
            full = (prefix_dir / inner).resolve()
            if _is_within_sandbox(full) and full.exists():
                return full

    return None


# ---------------------------------------------------------------------------
# 1. GET /api/download/advertiser-report -- Advertiser report as multi-sheet Excel
# ---------------------------------------------------------------------------
@router.get("/advertiser-report", dependencies=[Depends(require_paid)])
async def download_advertiser_report(
    advertiser_id: int = Query(..., description="Advertiser ID"),
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download comprehensive advertiser report as multi-sheet Excel.

    Sheets: 광고소재 목록, 채널별 통계, 광고비 추정
    """
    # Verify advertiser exists
    adv_result = await db.execute(
        select(Advertiser).where(Advertiser.id == advertiser_id)
    )
    advertiser = adv_result.scalar_one_or_none()
    if not advertiser:
        raise HTTPException(status_code=404, detail="Advertiser not found")

    date_from, date_to = _default_date_range(date_from, date_to)
    adv_name = advertiser.name or "unknown"

    wb = Workbook()

    # ── Sheet 1: 광고소재 목록 ──
    ws1 = wb.active
    ws1.title = "광고소재 목록"
    h1 = [
        "수집일시", "채널", "광고텍스트", "광고설명", "광고유형",
        "랜딩URL", "표시URL", "제품/서비스", "제품카테고리",
        "광고상품", "모델/셀럽", "포맷", "검증상태",
    ]
    ws1.append(h1)
    _style_header(ws1, len(h1))

    ads_query = (
        select(
            AdSnapshot.captured_at, AdSnapshot.channel,
            AdDetail.ad_text, AdDetail.ad_description, AdDetail.ad_type,
            AdDetail.url, AdDetail.display_url,
            AdDetail.product_name, AdDetail.product_category,
            AdDetail.ad_product_name, AdDetail.model_name,
            AdDetail.ad_format_type, AdDetail.verification_status,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id == advertiser_id)
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .order_by(AdSnapshot.captured_at.desc())
    )
    ads_result = await db.execute(ads_query)
    for r in ads_result.all():
        ws1.append([
            _kst_str(r[0]), _safe(r[1]), _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]), _safe(r[7]),
            _safe(r[8]), _safe(r[9]), _safe(r[10]), _safe(r[11]),
            _safe(r[12]),
        ])
    _auto_width(ws1, len(h1))
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: 채널별 통계 ──
    ws2 = wb.create_sheet("채널별 통계")
    h2 = ["채널", "광고 수", "최초발견", "최근발견", "총 추정 광고비"]
    ws2.append(h2)
    _style_header(ws2, len(h2))

    channel_stats_query = (
        select(
            AdSnapshot.channel,
            func.count(AdDetail.id).label("ad_count"),
            func.min(AdSnapshot.captured_at).label("first_seen"),
            func.max(AdSnapshot.captured_at).label("last_seen"),
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id == advertiser_id)
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .group_by(AdSnapshot.channel)
        .order_by(func.count(AdDetail.id).desc())
    )
    channel_stats = await db.execute(channel_stats_query)
    channel_stats_rows = channel_stats.all()

    # Get spend per channel
    channel_spend_query = (
        select(
            SpendEstimate.channel,
            func.sum(SpendEstimate.est_daily_spend).label("total_spend"),
        )
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .where(Campaign.advertiser_id == advertiser_id)
        .where(SpendEstimate.date >= date_from)
        .where(SpendEstimate.date <= date_to)
        .group_by(SpendEstimate.channel)
    )
    spend_result = await db.execute(channel_spend_query)
    spend_by_channel = {r[0]: r[1] for r in spend_result.all()}

    for r in channel_stats_rows:
        channel = _safe(r[0])
        spend = spend_by_channel.get(channel, 0)
        ws2.append([
            channel,
            r[1] or 0,
            _kst_str(r[2]),
            _kst_str(r[3]),
            round(spend, 0) if spend else 0,
        ])
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws2, len(h2))
    ws2.auto_filter.ref = ws2.dimensions

    # ── Sheet 3: 광고비 추정 ──
    ws3 = wb.create_sheet("광고비 추정")
    h3 = ["채널", "날짜", "일 추정 광고비", "신뢰도", "산출 방법"]
    ws3.append(h3)
    _style_header(ws3, len(h3))

    spend_detail_query = (
        select(
            SpendEstimate.channel, SpendEstimate.date,
            SpendEstimate.est_daily_spend, SpendEstimate.confidence,
            SpendEstimate.calculation_method,
        )
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .where(Campaign.advertiser_id == advertiser_id)
        .where(SpendEstimate.date >= date_from)
        .where(SpendEstimate.date <= date_to)
        .order_by(SpendEstimate.date.desc())
    )
    spend_detail_result = await db.execute(spend_detail_query)
    for r in spend_detail_result.all():
        ws3.append([
            _safe(r[0]),
            _kst_date_str(r[1]),
            round(r[2], 0) if r[2] else 0,
            round(r[3], 2) if r[3] else "",
            _safe(r[4]),
        ])
    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws3, len(h3))
    ws3.auto_filter.ref = ws3.dimensions

    # Return workbook
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = adv_name.replace(" ", "_").replace("/", "_")
    filename = f"advertiser_report_{safe_name}_{_today_str()}.xlsx"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_safe_content_disposition(filename),
    )


# ---------------------------------------------------------------------------
# 2. GET /api/download/advertiser-creatives -- ZIP of creative images
# ---------------------------------------------------------------------------
@router.get("/advertiser-creatives", dependencies=[Depends(require_paid), Depends(require_plan("full"))])
async def download_advertiser_creatives(
    advertiser_id: int = Query(..., description="Advertiser ID"),
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download all creative images for an advertiser as a ZIP file."""
    # Verify advertiser exists
    adv_result = await db.execute(
        select(Advertiser).where(Advertiser.id == advertiser_id)
    )
    advertiser = adv_result.scalar_one_or_none()
    if not advertiser:
        raise HTTPException(status_code=404, detail="Advertiser not found")

    date_from, date_to = _default_date_range(date_from, date_to)
    adv_name = advertiser.name or "unknown"

    # Fetch all creative image paths
    query = (
        select(
            AdDetail.id,
            AdDetail.creative_image_path,
            AdSnapshot.channel,
            AdSnapshot.captured_at,
            AdDetail.ad_text,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id == advertiser_id)
        .where(AdDetail.creative_image_path.isnot(None))
        .where(AdDetail.creative_image_path != "")
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .order_by(AdSnapshot.captured_at.desc())
    )
    result = await db.execute(query)
    rows = result.all()

    if not rows:
        raise HTTPException(status_code=404, detail="No creative images found for this advertiser")

    # Build ZIP in memory
    buf = io.BytesIO()
    added_count = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        seen_names = set()
        for r in rows:
            ad_id, img_path, channel, captured_at, ad_text = r
            resolved = _resolve_image_path(img_path)
            if resolved is None or not resolved.is_file():
                continue

            # Build a meaningful filename: channel/YYYY-MM-DD_adID.ext
            ext = resolved.suffix or ".webp"
            date_prefix = _kst_date_str(captured_at).replace("-", "") if captured_at else "unknown"
            arcname = f"{channel or 'unknown'}/{date_prefix}_{ad_id}{ext}"

            # Avoid duplicate names
            if arcname in seen_names:
                arcname = f"{channel or 'unknown'}/{date_prefix}_{ad_id}_dup{ext}"
            seen_names.add(arcname)

            try:
                zf.write(str(resolved), arcname)
                added_count += 1
            except Exception as e:
                logger.warning("Failed to add %s to ZIP: %s", resolved, e)

    if added_count == 0:
        raise HTTPException(
            status_code=404,
            detail="No accessible creative image files found on disk",
        )

    buf.seek(0)
    safe_name = adv_name.replace(" ", "_").replace("/", "_")
    filename = f"creatives_{safe_name}_{_today_str()}.zip"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/zip",
        headers=_safe_content_disposition(filename),
    )


# ---------------------------------------------------------------------------
# 3. GET /api/download/advertiser-list -- CSV of all advertisers with stats
# ---------------------------------------------------------------------------
@router.get("/advertiser-list", dependencies=[Depends(require_paid)])
async def download_advertiser_list(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download all advertisers as CSV with ad count, total spend, and channels."""
    # Sub-query: ad count per advertiser
    ad_count_sub = (
        select(
            AdDetail.advertiser_id,
            func.count(AdDetail.id).label("ad_count"),
        )
        .group_by(AdDetail.advertiser_id)
        .subquery()
    )

    # Sub-query: total spend per advertiser
    spend_sub = (
        select(
            Campaign.advertiser_id,
            func.sum(SpendEstimate.est_daily_spend).label("total_spend"),
        )
        .join(SpendEstimate, SpendEstimate.campaign_id == Campaign.id)
        .group_by(Campaign.advertiser_id)
        .subquery()
    )

    # Sub-query: distinct channels per advertiser
    channel_sub = (
        select(
            AdDetail.advertiser_id,
            func.group_concat(AdSnapshot.channel).label("channels"),
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id.isnot(None))
        .group_by(AdDetail.advertiser_id)
        .subquery()
    )

    query = (
        select(
            Advertiser.id,
            Advertiser.name,
            Industry.name.label("industry"),
            Advertiser.website,
            Advertiser.brand_name,
            ad_count_sub.c.ad_count,
            spend_sub.c.total_spend,
            channel_sub.c.channels,
            Advertiser.created_at,
        )
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .outerjoin(ad_count_sub, Advertiser.id == ad_count_sub.c.advertiser_id)
        .outerjoin(spend_sub, Advertiser.id == spend_sub.c.advertiser_id)
        .outerjoin(channel_sub, Advertiser.id == channel_sub.c.advertiser_id)
        .order_by(Advertiser.name)
    )

    result = await db.execute(query)
    rows_raw = result.all()

    header = [
        "ID", "광고주명", "업종", "웹사이트", "브랜드명",
        "광고 수", "총 추정 광고비(원)", "채널", "등록일",
    ]
    rows = []
    for r in rows_raw:
        rows.append([
            _safe(r[0]),
            _safe(r[1]),
            _safe(r[2]),
            _safe(r[3]),
            _safe(r[4]),
            _safe(r[5] or 0),
            _safe(round(r[6], 0) if r[6] else 0),
            _safe(",".join(sorted(set(c.strip() for c in (r[7] or "").split(",") if c.strip())))),
            _kst_str(r[8]),
        ])

    buf = io.StringIO()
    buf.write(UTF8_BOM)
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    buf.seek(0)

    filename = f"adscope_advertiser_list_{_today_str()}.csv"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers=_safe_content_disposition(filename),
    )


# ---------------------------------------------------------------------------
# 4. GET /api/download/gallery-selection -- ZIP of selected gallery images
# ---------------------------------------------------------------------------
@router.get("/gallery-selection", dependencies=[Depends(require_paid), Depends(require_plan("full"))])
async def download_gallery_selection(
    ids: str = Query(..., description="Comma-separated ad_detail IDs"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Download selected gallery creative images as a ZIP file.

    Pass ad_detail IDs as comma-separated values: ?ids=1,2,3
    """
    # Parse IDs
    try:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid IDs format. Use comma-separated integers.")

    if not id_list:
        raise HTTPException(status_code=400, detail="No IDs provided")

    if len(id_list) > 500:
        raise HTTPException(status_code=400, detail="Maximum 500 items per download")

    # Fetch image paths for the selected IDs
    query = (
        select(
            AdDetail.id,
            AdDetail.creative_image_path,
            AdDetail.advertiser_name_raw,
            AdSnapshot.channel,
            AdSnapshot.captured_at,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.id.in_(id_list))
        .where(AdDetail.creative_image_path.isnot(None))
        .where(AdDetail.creative_image_path != "")
    )
    result = await db.execute(query)
    rows = result.all()

    if not rows:
        raise HTTPException(status_code=404, detail="No creative images found for the selected items")

    # Build ZIP
    buf = io.BytesIO()
    added_count = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        seen_names = set()
        for r in rows:
            ad_id, img_path, adv_name, channel, captured_at = r
            resolved = _resolve_image_path(img_path)
            if resolved is None or not resolved.is_file():
                continue

            ext = resolved.suffix or ".webp"
            date_prefix = _kst_date_str(captured_at).replace("-", "") if captured_at else "unknown"
            safe_adv = (adv_name or "unknown").replace(" ", "_").replace("/", "_")[:30]
            arcname = f"{safe_adv}_{channel or 'unknown'}_{date_prefix}_{ad_id}{ext}"

            if arcname in seen_names:
                arcname = f"{safe_adv}_{channel or 'unknown'}_{date_prefix}_{ad_id}_dup{ext}"
            seen_names.add(arcname)

            try:
                zf.write(str(resolved), arcname)
                added_count += 1
            except Exception as e:
                logger.warning("Failed to add %s to ZIP: %s", resolved, e)

    if added_count == 0:
        raise HTTPException(
            status_code=404,
            detail="No accessible creative image files found on disk",
        )

    buf.seek(0)
    filename = f"gallery_selection_{_today_str()}.zip"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/zip",
        headers=_safe_content_disposition(filename),
    )
