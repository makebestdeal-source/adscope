"""Export API -- CSV & Excel for gallery, advertisers, spend, report, social."""

import csv
import hashlib
import io
import os
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import false, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from api.deps import get_current_user, require_paid, require_plan
from database import get_db
from database.models import (
    AdDetail,
    AdSnapshot,
    Advertiser,
    BrandChannelContent,
    Campaign,
    ChannelStats,
    Industry,
    SmartStoreSnapshot,
    SmartStoreTrackedProduct,
    SpendEstimate,
    User,
)
from processor.channel_utils import (
    GALLERY_EXCLUDED_CHANNELS,
    get_gallery_ad_channels,
    normalize_channel_for_display,
)

router = APIRouter(
    prefix="/api/export",
    tags=["export"],
    dependencies=[Depends(require_paid)],
)

KST = timezone(timedelta(hours=9))
CACHE_DIR = Path("cache/exports")
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# ResearchAD 18컬럼 표준 포맷 — 채널 매핑 상수
# ---------------------------------------------------------------------------
# channel → (매체, 카테고리, 카테고리세분, 앱/매체, 매체카테고리/플랫폼, 플랫폼)
_CHANNEL_MAP: dict[str, tuple[str, str, str, str, str, str]] = {
    "naver_search":     ("포털", "포털", "포털", "네이버 검색",         "포털(대)검색(중)",         "-"),
    "naver_da":         ("포털", "포털", "포털", "네이버 DA",            "포털(대)디스플레이(중)",   "-"),
    "naver_shopping":   ("포털", "포털", "포털", "네이버 쇼핑",          "포털(대)쇼핑(중)",         "-"),
    "kakao_da":         ("포털", "포털", "포털", "카카오 DA",            "포털(대)디스플레이(중)",   "-"),
    "youtube_ads":      ("SNS",  "SNS",  "SNS",  "유튜브",               "SNS(대)동영상(중)",        "-"),
    "meta":             ("SNS",  "SNS",  "SNS",  "메타(페이스북/인스타)", "SNS(대)소셜(중)",          "-"),
    "tiktok_ads":       ("SNS",  "SNS",  "SNS",  "틱톡",                 "SNS(대)숏폼(중)",          "-"),
    "google_gdn":       ("네트워크", "네트워크", "GDN",    "구글 GDN",   "네트워크(대)디스플레이(중)", "GDN"),
    "google_search_ads":("네트워크", "네트워크", "구글",   "구글 검색광고", "네트워크(대)검색(중)",   "-"),
}
_CHANNEL_MAP_DEFAULT = ("기타", "기타", "기타", "", "", "-")

_PLACEMENT_MAP: dict[str, str] = {
    "naver_main":            "메인",
    "naver_news":            "뉴스",
    "naver_content":         "콘텐츠",
    "naver_shopping_search": "쇼핑검색",
    "meta_ads_library":      "피드",
    "google_ads_transparency": "-",
    "kakao_main":            "메인",
    "youtube_preroll":       "프리롤",
    "tiktok_top_ads":        "피드",
    "tiktok_feed":           "피드",
    "google_search":         "검색결과",
    "naver_powerlink":       "파워링크",
    "naver_bizsite":         "비즈사이트",
}


def _ch(channel: str) -> tuple[str, str, str, str, str, str]:
    return _CHANNEL_MAP.get(channel, _CHANNEL_MAP_DEFAULT)


def _rad_date(dt: datetime | None) -> str:
    """ResearchAD 날짜 포맷 YY.MM.DD"""
    if dt is None:
        return ""
    kst_dt = dt + timedelta(hours=9)
    return kst_dt.strftime("%y.%m.%d")


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

UTF8_BOM = "\ufeff"


def _safe(value) -> str:
    if value is None:
        return ""
    return str(value)


def _kst_str(dt: datetime | None) -> str:
    if dt is None:
        return ""
    kst_dt = dt + timedelta(hours=9)
    return kst_dt.strftime("%Y-%m-%d %H:%M:%S")


def _kst_date_str(dt: datetime | None) -> str:
    if dt is None:
        return ""
    kst_dt = dt + timedelta(hours=9)
    return kst_dt.strftime("%Y-%m-%d")


def _today_str() -> str:
    return datetime.now(KST).strftime("%Y%m%d")


def _csv_streaming_response(filename: str, header: list[str], rows: list[list[str]]) -> StreamingResponse:
    buf = io.StringIO()
    buf.write(UTF8_BOM)
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers=_safe_cd(filename),
    )


def _default_date_range(date_from: datetime | None, date_to: datetime | None):
    if date_to is None:
        date_to = datetime.utcnow()
    if date_from is None:
        date_from = date_to - timedelta(days=30)
    return date_from, date_to


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _safe_cd(filename: str) -> dict[str, str]:
    """RFC 5987 Content-Disposition with non-ASCII support."""
    ascii_name = filename.encode("ascii", "replace").decode("ascii").replace("?", "_")
    utf8_name = quote(filename, safe="")
    return {"Content-Disposition": f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"}


def _xlsx_response(wb: Workbook, filename: str, cache_key: str | None = None) -> FileResponse | StreamingResponse:
    """Save workbook and return as response. Uses cache if cache_key provided."""
    if cache_key:
        cache_path = CACHE_DIR / f"{cache_key}.xlsx"
        # Return cached file if exists and less than 1 hour old
        if cache_path.exists():
            age = datetime.now().timestamp() - cache_path.stat().st_mtime
            if age < 3600:
                return FileResponse(
                    str(cache_path),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=_safe_cd(filename),
                )
        wb.save(str(cache_path))
        return FileResponse(
            str(cache_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_safe_cd(filename),
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_safe_cd(filename),
    )


def _style_header(ws, col_count: int):
    """Apply header styling to first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_width(ws, col_count: int, max_width: int = 40):
    """Auto-adjust column widths based on content."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 10)


def _cache_key(*parts) -> str:
    """Generate cache key from query parameters."""
    raw = "|".join(str(p) for p in parts)
    return hashlib.md5(raw.encode()).hexdigest()[:12]


# ---------------------------------------------------------------------------
# 1a. GET /api/export/gallery
# ---------------------------------------------------------------------------
@router.get("/gallery", dependencies=[Depends(require_plan("full"))])
async def export_gallery(
    channel: str | None = None,
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export gallery (ad details) as CSV."""
    date_from, date_to = _default_date_range(date_from, date_to)

    query = (
        select(
            AdSnapshot.captured_at,
            AdSnapshot.channel,
            AdDetail.advertiser_name_raw,
            AdDetail.ad_text,
            AdDetail.ad_type,
            AdDetail.url,
            AdDetail.product_category,
            AdDetail.verification_status,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .where(~AdSnapshot.channel.in_(GALLERY_EXCLUDED_CHANNELS))
        .order_by(AdSnapshot.captured_at.desc())
    )

    if channel:
        ad_channels = get_gallery_ad_channels(channel)
        if not ad_channels:
            query = query.where(false())
        elif len(ad_channels) == 1:
            query = query.where(AdSnapshot.channel == next(iter(ad_channels)))
        else:
            query = query.where(AdSnapshot.channel.in_(ad_channels))
    if advertiser_id:
        query = query.where(AdDetail.advertiser_id == advertiser_id)

    result = await db.execute(query)
    rows_raw = result.all()

    header = [
        "captured_at", "channel", "advertiser_name_raw", "ad_text",
        "ad_type", "url", "product_category", "verification_status",
    ]
    rows = []
    for r in rows_raw:
        rows.append([
            _kst_str(r[0]),
            _safe(normalize_channel_for_display(r[1])),
            _safe(r[2]),
            _safe(r[3]),
            _safe(r[4]),
            _safe(r[5]),
            _safe(r[6]),
            _safe(r[7]),
        ])

    filename = f"adscope_gallery_{_today_str()}.csv"
    return _csv_streaming_response(filename, header, rows)


# ---------------------------------------------------------------------------
# 1b. GET /api/export/advertisers
# ---------------------------------------------------------------------------
@router.get("/advertisers")
async def export_advertisers(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export advertiser list as CSV with ad_count, total_spend, channels."""
    # Sub-query: ad count per advertiser
    ad_count_sub = (
        select(
            AdDetail.advertiser_id,
            func.count(AdDetail.id).label("ad_count"),
        )
        .group_by(AdDetail.advertiser_id)
        .subquery()
    )

    # Sub-query: total spend per advertiser (via campaign)
    spend_sub = (
        select(
            Campaign.advertiser_id,
            func.sum(SpendEstimate.est_daily_spend).label("total_spend"),
        )
        .join(SpendEstimate, SpendEstimate.campaign_id == Campaign.id)
        .group_by(Campaign.advertiser_id)
        .subquery()
    )

    # Sub-query: distinct channels per advertiser
    channel_sub = (
        select(
            AdDetail.advertiser_id,
            func.group_concat(AdSnapshot.channel.distinct()).label("channels"),
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id.isnot(None))
        .group_by(AdDetail.advertiser_id)
        .subquery()
    )

    query = (
        select(
            Advertiser.name,
            Industry.name.label("industry"),
            Advertiser.website,
            ad_count_sub.c.ad_count,
            spend_sub.c.total_spend,
            channel_sub.c.channels,
        )
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .outerjoin(ad_count_sub, Advertiser.id == ad_count_sub.c.advertiser_id)
        .outerjoin(spend_sub, Advertiser.id == spend_sub.c.advertiser_id)
        .outerjoin(channel_sub, Advertiser.id == channel_sub.c.advertiser_id)
        .order_by(Advertiser.name)
    )

    result = await db.execute(query)
    rows_raw = result.all()

    header = ["name", "industry", "website", "ad_count", "total_spend", "channels"]
    rows = []
    for r in rows_raw:
        rows.append([
            _safe(r[0]),
            _safe(r[1]),
            _safe(r[2]),
            _safe(r[3] or 0),
            _safe(round(r[4], 0) if r[4] else 0),
            _safe(r[5]),
        ])

    filename = f"adscope_advertisers_{_today_str()}.csv"
    return _csv_streaming_response(filename, header, rows)


# ---------------------------------------------------------------------------
# 1c. GET /api/export/spend
# ---------------------------------------------------------------------------
@router.get("/spend")
async def export_spend(
    advertiser_id: int | None = None,
    channel: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export spend estimates as CSV."""
    date_from, date_to = _default_date_range(date_from, date_to)

    query = (
        select(
            Advertiser.name.label("advertiser_name"),
            SpendEstimate.channel,
            SpendEstimate.est_daily_spend,
            SpendEstimate.confidence,
            SpendEstimate.calculation_method,
            SpendEstimate.date,
        )
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .join(Advertiser, Campaign.advertiser_id == Advertiser.id)
        .where(SpendEstimate.date >= date_from)
        .where(SpendEstimate.date <= date_to)
        .order_by(SpendEstimate.date.desc())
    )

    if advertiser_id:
        query = query.where(Campaign.advertiser_id == advertiser_id)
    if channel:
        query = query.where(SpendEstimate.channel == channel)

    result = await db.execute(query)
    rows_raw = result.all()

    header = [
        "advertiser_name", "channel", "est_daily_spend",
        "confidence_score", "calculation_method", "date",
    ]
    rows = []
    for r in rows_raw:
        rows.append([
            _safe(r[0]),
            _safe(r[1]),
            _safe(round(r[2], 0) if r[2] else 0),
            _safe(round(r[3], 2) if r[3] else ""),
            _safe(r[4]),
            _kst_date_str(r[5]),
        ])

    filename = f"adscope_spend_{_today_str()}.csv"
    return _csv_streaming_response(filename, header, rows)


# ---------------------------------------------------------------------------
# 1d. GET /api/export/report/{advertiser_id}
# ---------------------------------------------------------------------------
@router.get("/report/{advertiser_id}")
async def export_report(
    advertiser_id: int,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export comprehensive report for a single advertiser (ads + campaigns + spend)."""
    # Verify advertiser exists
    adv_result = await db.execute(
        select(Advertiser).where(Advertiser.id == advertiser_id)
    )
    advertiser = adv_result.scalar_one_or_none()
    if not advertiser:
        raise HTTPException(status_code=404, detail="Advertiser not found")

    date_from, date_to = _default_date_range(date_from, date_to)

    # --- Section 1: Ad creatives (6W 확장) ---
    ads_query = (
        select(
            AdSnapshot.captured_at,
            AdSnapshot.channel,
            AdDetail.ad_text,
            AdDetail.ad_description,
            AdDetail.ad_type,
            AdDetail.url,
            AdDetail.display_url,
            AdDetail.product_name,
            AdDetail.product_category,
            AdDetail.ad_product_name,
            AdDetail.model_name,
            AdDetail.ad_format_type,
            AdDetail.verification_status,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id == advertiser_id)
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .order_by(AdSnapshot.captured_at.desc())
    )
    ads_result = await db.execute(ads_query)
    ads_rows = ads_result.all()

    # --- Section 2: Campaigns ---
    camp_query = (
        select(
            Campaign.campaign_name,
            Campaign.channel,
            Campaign.first_seen,
            Campaign.last_seen,
            Campaign.is_active,
            Campaign.total_est_spend,
            Campaign.snapshot_count,
            Campaign.product_service,
            Campaign.model_info,
        )
        .where(Campaign.advertiser_id == advertiser_id)
        .order_by(Campaign.last_seen.desc())
    )
    camp_result = await db.execute(camp_query)
    camp_rows = camp_result.all()

    # --- Section 3: Spend estimates ---
    spend_query = (
        select(
            SpendEstimate.channel,
            SpendEstimate.date,
            SpendEstimate.est_daily_spend,
            SpendEstimate.confidence,
            SpendEstimate.calculation_method,
        )
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .where(Campaign.advertiser_id == advertiser_id)
        .where(SpendEstimate.date >= date_from)
        .where(SpendEstimate.date <= date_to)
        .order_by(SpendEstimate.date.desc())
    )
    spend_result = await db.execute(spend_query)
    spend_rows = spend_result.all()

    # --- Build unified CSV ---
    buf = io.StringIO()
    buf.write(UTF8_BOM)
    writer = csv.writer(buf)

    # Header info
    adv_name = advertiser.name or ""
    writer.writerow([f"AdScope Report - {adv_name}"])
    writer.writerow([f"Period: {_kst_date_str(date_from)} ~ {_kst_date_str(date_to)}"])
    writer.writerow([f"Generated: {datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')} KST"])
    writer.writerow([])

    # Section 1: Ads
    writer.writerow(["=== Ad Creatives ==="])
    writer.writerow([
        "수집일시", "채널", "광고텍스트", "광고설명", "광고유형",
        "랜딩URL", "표시URL", "제품/서비스", "제품카테고리",
        "광고상품", "모델/셀럽", "포맷", "검증상태",
    ])
    for r in ads_rows:
        writer.writerow([
            _kst_str(r[0]), _safe(r[1]), _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]), _safe(r[7]),
            _safe(r[8]), _safe(r[9]), _safe(r[10]), _safe(r[11]),
            _safe(r[12]),
        ])
    writer.writerow([])

    # Section 2: Campaigns
    writer.writerow(["=== Campaigns ==="])
    writer.writerow([
        "캠페인명", "채널", "최초발견", "최근발견", "활성여부",
        "총추정광고비", "스냅샷수", "제품/서비스", "모델",
    ])
    for r in camp_rows:
        writer.writerow([
            _safe(r[0]), _safe(r[1]), _kst_str(r[2]), _kst_str(r[3]),
            "Y" if r[4] else "N", _safe(round(r[5], 0) if r[5] else 0),
            _safe(r[6]), _safe(r[7]), _safe(r[8]),
        ])
    writer.writerow([])

    # Section 3: Spend Estimates
    writer.writerow(["=== Spend Estimates ==="])
    writer.writerow([
        "channel", "date", "est_daily_spend",
        "confidence_score", "calculation_method",
    ])
    for r in spend_rows:
        writer.writerow([
            _safe(r[0]), _kst_date_str(r[1]),
            _safe(round(r[2], 0) if r[2] else 0),
            _safe(round(r[3], 2) if r[3] else ""),
            _safe(r[4]),
        ])

    buf.seek(0)
    safe_name = adv_name.replace(" ", "_").replace("/", "_")
    filename = f"adscope_report_{safe_name}_{_today_str()}.csv"

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers=_safe_cd(filename),
    )


# ---------------------------------------------------------------------------
# 2. GET /api/export/social -- Social content CSV
# ---------------------------------------------------------------------------
@router.get("/social")
async def export_social(
    channel: str | None = None,
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export social content (brand_channel_contents) as CSV."""
    date_from, date_to = _default_date_range(date_from, date_to)

    query = (
        select(
            BrandChannelContent.discovered_at,
            BrandChannelContent.platform,
            Advertiser.name,
            BrandChannelContent.title,
            BrandChannelContent.content_type,
            BrandChannelContent.view_count,
            BrandChannelContent.like_count,
            BrandChannelContent.upload_date,
            BrandChannelContent.content_id,
        )
        .join(Advertiser, BrandChannelContent.advertiser_id == Advertiser.id)
        .where(BrandChannelContent.discovered_at >= date_from)
        .where(BrandChannelContent.discovered_at <= date_to)
        .order_by(BrandChannelContent.discovered_at.desc())
    )

    if channel and channel in ("youtube", "instagram", "meta"):
        if channel == "meta":
            query = query.where(BrandChannelContent.platform.in_(["meta", "instagram", "facebook"]))
        else:
            query = query.where(BrandChannelContent.platform == channel)
    if advertiser_id:
        query = query.where(BrandChannelContent.advertiser_id == advertiser_id)

    result = await db.execute(query)
    rows_raw = result.all()

    header = [
        "discovered_at", "platform", "advertiser", "title",
        "content_type", "view_count", "like_count", "upload_date", "content_url",
    ]
    rows = []
    for r in rows_raw:
        platform = _safe(r[1])
        cid = _safe(r[8])
        if platform == "youtube" and cid:
            url = f"https://www.youtube.com/watch?v={cid}"
        elif platform in ("instagram", "meta") and cid:
            url = f"https://www.instagram.com/p/{cid}/"
        else:
            url = ""
        rows.append([
            _kst_str(r[0]), platform, _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5] or 0), _safe(r[6] or 0),
            _kst_date_str(r[7]), url,
        ])

    filename = f"adscope_social_{_today_str()}.csv"
    return _csv_streaming_response(filename, header, rows)


# ===========================================================================
# Excel Export Endpoints
# ===========================================================================

# ---------------------------------------------------------------------------
# 3a. GET /api/export/gallery.xlsx
# ---------------------------------------------------------------------------
@router.get("/gallery.xlsx", dependencies=[Depends(require_plan("full"))])
async def export_gallery_xlsx(
    channel: str | None = None,
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export gallery as Excel (.xlsx)."""
    date_from, date_to = _default_date_range(date_from, date_to)
    ck = _cache_key("gallery", channel, advertiser_id, _kst_date_str(date_from), _kst_date_str(date_to))

    query = (
        select(
            AdSnapshot.captured_at,
            AdSnapshot.channel,
            AdDetail.advertiser_name_raw,
            AdDetail.ad_text,
            AdDetail.ad_type,
            AdDetail.url,
            AdDetail.product_category,
            AdDetail.verification_status,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .where(~AdSnapshot.channel.in_(GALLERY_EXCLUDED_CHANNELS))
        .order_by(AdSnapshot.captured_at.desc())
    )
    if channel:
        ad_channels = get_gallery_ad_channels(channel)
        if not ad_channels:
            query = query.where(false())
        elif len(ad_channels) == 1:
            query = query.where(AdSnapshot.channel == next(iter(ad_channels)))
        else:
            query = query.where(AdSnapshot.channel.in_(ad_channels))
    if advertiser_id:
        query = query.where(AdDetail.advertiser_id == advertiser_id)

    result = await db.execute(query)
    rows_raw = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ad Creatives"

    headers = ["수집일시", "채널", "광고주", "광고 텍스트", "광고 유형", "URL", "제품 카테고리", "검증"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows_raw:
        ws.append([
            _kst_str(r[0]), _safe(normalize_channel_for_display(r[1])), _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]), _safe(r[7]),
        ])

    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions
    filename = f"adscope_gallery_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 3b. GET /api/export/advertisers.xlsx
# ---------------------------------------------------------------------------
@router.get("/advertisers.xlsx")
async def export_advertisers_xlsx(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export advertiser list as Excel."""
    ck = _cache_key("advertisers", _today_str())

    ad_count_sub = (
        select(AdDetail.advertiser_id, func.count(AdDetail.id).label("ad_count"))
        .group_by(AdDetail.advertiser_id).subquery()
    )
    spend_sub = (
        select(Campaign.advertiser_id, func.sum(SpendEstimate.est_daily_spend).label("total_spend"))
        .join(SpendEstimate, SpendEstimate.campaign_id == Campaign.id)
        .group_by(Campaign.advertiser_id).subquery()
    )
    channel_sub = (
        select(AdDetail.advertiser_id, func.group_concat(AdSnapshot.channel.distinct()).label("channels"))
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id.isnot(None))
        .group_by(AdDetail.advertiser_id).subquery()
    )

    query = (
        select(
            Advertiser.name, Industry.name.label("industry"), Advertiser.website,
            ad_count_sub.c.ad_count, spend_sub.c.total_spend, channel_sub.c.channels,
        )
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .outerjoin(ad_count_sub, Advertiser.id == ad_count_sub.c.advertiser_id)
        .outerjoin(spend_sub, Advertiser.id == spend_sub.c.advertiser_id)
        .outerjoin(channel_sub, Advertiser.id == channel_sub.c.advertiser_id)
        .order_by(Advertiser.name)
    )

    result = await db.execute(query)
    rows_raw = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Advertisers"

    headers = ["광고주명", "업종", "웹사이트", "광고 수", "총 추정 광고비", "채널"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows_raw:
        ws.append([
            _safe(r[0]), _safe(r[1]), _safe(r[2]),
            r[3] or 0, round(r[4], 0) if r[4] else 0, _safe(r[5]),
        ])

    # Number format for spend column
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0'

    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions
    filename = f"adscope_advertisers_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 3c. GET /api/export/spend.xlsx
# ---------------------------------------------------------------------------
@router.get("/spend.xlsx")
async def export_spend_xlsx(
    advertiser_id: int | None = None,
    channel: str | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export spend estimates as Excel."""
    date_from, date_to = _default_date_range(date_from, date_to)
    ck = _cache_key("spend", advertiser_id, channel, _kst_date_str(date_from), _kst_date_str(date_to))

    query = (
        select(
            Advertiser.name.label("advertiser_name"),
            SpendEstimate.channel, SpendEstimate.est_daily_spend,
            SpendEstimate.confidence, SpendEstimate.calculation_method, SpendEstimate.date,
        )
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .join(Advertiser, Campaign.advertiser_id == Advertiser.id)
        .where(SpendEstimate.date >= date_from)
        .where(SpendEstimate.date <= date_to)
        .order_by(SpendEstimate.date.desc())
    )
    if advertiser_id:
        query = query.where(Campaign.advertiser_id == advertiser_id)
    if channel:
        query = query.where(SpendEstimate.channel == channel)

    result = await db.execute(query)
    rows_raw = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Spend"

    headers = ["광고주", "채널", "일 추정 광고비", "신뢰도", "산출 방법", "날짜"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows_raw:
        ws.append([
            _safe(r[0]), _safe(r[1]),
            round(r[2], 0) if r[2] else 0,
            round(r[3], 2) if r[3] else "",
            _safe(r[4]), _kst_date_str(r[5]),
        ])

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'

    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions
    filename = f"adscope_spend_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 3d. GET /api/export/report/{advertiser_id}.xlsx
# ---------------------------------------------------------------------------
@router.get("/report/{advertiser_id}.xlsx")
async def export_report_xlsx(
    advertiser_id: int,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export comprehensive report as Excel with multiple sheets."""
    adv_result = await db.execute(select(Advertiser).where(Advertiser.id == advertiser_id))
    advertiser = adv_result.scalar_one_or_none()
    if not advertiser:
        raise HTTPException(status_code=404, detail="Advertiser not found")

    date_from, date_to = _default_date_range(date_from, date_to)
    adv_name = advertiser.name or ""
    ck = _cache_key("report", advertiser_id, _kst_date_str(date_from), _kst_date_str(date_to))

    wb = Workbook()

    # --- Sheet 1: Ad Creatives (6W 확장) ---
    ws1 = wb.active
    ws1.title = "광고소재"
    h1 = ["수집일시", "채널", "광고텍스트", "광고설명", "광고유형",
           "랜딩URL", "표시URL", "제품/서비스", "제품카테고리",
           "광고상품", "모델/셀럽", "포맷", "검증상태"]
    ws1.append(h1)
    _style_header(ws1, len(h1))

    ads_result = await db.execute(
        select(
            AdSnapshot.captured_at, AdSnapshot.channel,
            AdDetail.ad_text, AdDetail.ad_description, AdDetail.ad_type,
            AdDetail.url, AdDetail.display_url,
            AdDetail.product_name, AdDetail.product_category,
            AdDetail.ad_product_name, AdDetail.model_name,
            AdDetail.ad_format_type, AdDetail.verification_status,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .where(AdDetail.advertiser_id == advertiser_id)
        .where(AdSnapshot.captured_at >= date_from)
        .where(AdSnapshot.captured_at <= date_to)
        .order_by(AdSnapshot.captured_at.desc())
    )
    for r in ads_result.all():
        ws1.append([
            _kst_str(r[0]), _safe(r[1]), _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]), _safe(r[7]),
            _safe(r[8]), _safe(r[9]), _safe(r[10]), _safe(r[11]),
            _safe(r[12]),
        ])
    _auto_width(ws1, len(h1))
    ws1.auto_filter.ref = ws1.dimensions

    # --- Sheet 2: Campaigns ---
    ws2 = wb.create_sheet("캠페인")
    h2 = ["캠페인명", "채널", "최초발견", "최근발견", "활성여부",
           "총추정광고비", "스냅샷수", "제품/서비스", "모델"]
    ws2.append(h2)
    _style_header(ws2, len(h2))

    camp_result = await db.execute(
        select(Campaign.campaign_name, Campaign.channel, Campaign.first_seen, Campaign.last_seen,
               Campaign.is_active, Campaign.total_est_spend, Campaign.snapshot_count,
               Campaign.product_service, Campaign.model_info)
        .where(Campaign.advertiser_id == advertiser_id)
        .order_by(Campaign.last_seen.desc())
    )
    for r in camp_result.all():
        ws2.append([
            _safe(r[0]), _safe(r[1]), _kst_str(r[2]), _kst_str(r[3]),
            "Y" if r[4] else "N", round(r[5], 0) if r[5] else 0, r[6] or 0,
            _safe(r[7]), _safe(r[8]),
        ])
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws2, len(h2))

    # --- Sheet 3: Spend ---
    ws3 = wb.create_sheet("Spend")
    h3 = ["채널", "날짜", "일 추정 광고비", "신뢰도", "산출 방법"]
    ws3.append(h3)
    _style_header(ws3, len(h3))

    spend_result = await db.execute(
        select(SpendEstimate.channel, SpendEstimate.date,
               SpendEstimate.est_daily_spend, SpendEstimate.confidence, SpendEstimate.calculation_method)
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .where(Campaign.advertiser_id == advertiser_id)
        .where(SpendEstimate.date >= date_from)
        .where(SpendEstimate.date <= date_to)
        .order_by(SpendEstimate.date.desc())
    )
    for r in spend_result.all():
        ws3.append([
            _safe(r[0]), _kst_date_str(r[1]),
            round(r[2], 0) if r[2] else 0,
            round(r[3], 2) if r[3] else "", _safe(r[4]),
        ])
    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws3, len(h3))

    # --- Sheet 4: Social Content ---
    ws4 = wb.create_sheet("Social")
    h4 = ["플랫폼", "게시일", "제목", "유형", "조회수", "좋아요", "URL"]
    ws4.append(h4)
    _style_header(ws4, len(h4))

    social_result = await db.execute(
        select(
            BrandChannelContent.platform, BrandChannelContent.upload_date,
            BrandChannelContent.title, BrandChannelContent.content_type,
            BrandChannelContent.view_count, BrandChannelContent.like_count,
            BrandChannelContent.content_id,
        )
        .where(BrandChannelContent.advertiser_id == advertiser_id)
        .order_by(BrandChannelContent.upload_date.desc())
    )
    for r in social_result.all():
        platform = _safe(r[0])
        cid = _safe(r[6])
        if platform == "youtube" and cid:
            url = f"https://www.youtube.com/watch?v={cid}"
        elif platform == "instagram" and cid:
            url = f"https://www.instagram.com/p/{cid}/"
        else:
            url = ""
        ws4.append([
            platform, _kst_date_str(r[1]), _safe(r[2]), _safe(r[3]),
            r[4] or 0, r[5] or 0, url,
        ])
    for row in ws4.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws4, len(h4))

    safe_name = adv_name.replace(" ", "_").replace("/", "_")
    filename = f"adscope_report_{safe_name}_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 3e. GET /api/export/social.xlsx
# ---------------------------------------------------------------------------
@router.get("/social.xlsx")
async def export_social_xlsx(
    channel: str | None = None,
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export social content as Excel."""
    date_from, date_to = _default_date_range(date_from, date_to)
    ck = _cache_key("social", channel, advertiser_id, _kst_date_str(date_from), _kst_date_str(date_to))

    query = (
        select(
            BrandChannelContent.discovered_at, BrandChannelContent.platform,
            Advertiser.name, BrandChannelContent.title, BrandChannelContent.content_type,
            BrandChannelContent.view_count, BrandChannelContent.like_count,
            BrandChannelContent.upload_date, BrandChannelContent.content_id,
        )
        .join(Advertiser, BrandChannelContent.advertiser_id == Advertiser.id)
        .where(BrandChannelContent.discovered_at >= date_from)
        .where(BrandChannelContent.discovered_at <= date_to)
        .order_by(BrandChannelContent.discovered_at.desc())
    )
    if channel and channel in ("youtube", "instagram", "meta"):
        if channel == "meta":
            query = query.where(BrandChannelContent.platform.in_(["meta", "instagram", "facebook"]))
        else:
            query = query.where(BrandChannelContent.platform == channel)
    if advertiser_id:
        query = query.where(BrandChannelContent.advertiser_id == advertiser_id)

    result = await db.execute(query)
    rows_raw = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Social Content"

    headers = ["수집일시", "플랫폼", "광고주", "제목", "유형", "조회수", "좋아요", "게시일", "URL"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows_raw:
        platform = _safe(r[1])
        cid = _safe(r[8])
        if platform == "youtube" and cid:
            url = f"https://www.youtube.com/watch?v={cid}"
        elif platform == "instagram" and cid:
            url = f"https://www.instagram.com/p/{cid}/"
        else:
            url = ""
        ws.append([
            _kst_str(r[0]), platform, _safe(r[2]), _safe(r[3]),
            _safe(r[4]), r[5] or 0, r[6] or 0, _kst_date_str(r[7]), url,
        ])

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = '#,##0'

    _auto_width(ws, len(headers))
    ws.auto_filter.ref = ws.dimensions
    filename = f"adscope_social_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 4a. GET /api/export/ad-scope.xlsx -- Ad Scope raw data (multi-sheet)
# ---------------------------------------------------------------------------
@router.get("/ad-scope.xlsx")
async def export_ad_scope_xlsx(
    channel: str | None = None,
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Ad Scope raw data -- 4 sheets: creatives, advertisers, campaigns, spend."""
    date_from, date_to = _default_date_range(date_from, date_to)
    ck = _cache_key("ad-scope", channel, advertiser_id,
                     _kst_date_str(date_from), _kst_date_str(date_to))

    wb = Workbook()

    # --- Sheet 1: 광고소재 (ResearchAD 18컬럼 표준) ---
    ws1 = wb.active
    ws1.title = "광고소재"
    h1 = ["매체", "카테고리", "카테고리 세분", "구분",
          "광고주", "브랜드/제품",
          "업종(대)", "업종(중)", "업종(소)",
          "앱/매체", "매체 카테고리/플랫폼", "게재면", "플랫폼",
          "연", "월", "시작일", "종료일", "광고비(전원)"]
    ws1.append(h1)
    _style_header(ws1, len(h1))

    # 광고비 집계: advertiser_id + channel 단위 합산 (날짜 범위 내)
    spend_sub = (
        select(
            Campaign.advertiser_id,
            Campaign.channel,
            func.sum(SpendEstimate.est_daily_spend).label("total_spend"),
        )
        .join(SpendEstimate, SpendEstimate.campaign_id == Campaign.id)
        .where(SpendEstimate.date >= date_from, SpendEstimate.date <= date_to)
        .group_by(Campaign.advertiser_id, Campaign.channel)
        .subquery()
    )

    q1 = (
        select(
            AdSnapshot.channel,
            AdSnapshot.captured_at,
            AdDetail.advertiser_name_raw,
            AdDetail.brand,
            Industry.industry_large,
            Industry.industry_medium,
            Industry.name,
            AdDetail.ad_placement,
            AdDetail.retargeting_network,
            AdDetail.first_seen_at,
            AdDetail.last_seen_at,
            spend_sub.c.total_spend,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .outerjoin(Advertiser, AdDetail.advertiser_id == Advertiser.id)
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .outerjoin(
            spend_sub,
            (spend_sub.c.advertiser_id == AdDetail.advertiser_id)
            & (spend_sub.c.channel == AdSnapshot.channel),
        )
        .where(AdSnapshot.captured_at >= date_from, AdSnapshot.captured_at <= date_to)
        .order_by(AdSnapshot.captured_at.desc())
    )
    if channel:
        q1 = q1.where(AdSnapshot.channel == channel)
    if advertiser_id:
        q1 = q1.where(AdDetail.advertiser_id == advertiser_id)

    for r in (await db.execute(q1)).all():
        ch = _safe(r[0])
        media, cat, cat_detail, app_media, media_platform, platform = _ch(ch)
        placement = _PLACEMENT_MAP.get(_safe(r[7]), _safe(r[7]))
        if r[8]:  # retargeting_network 있으면 플랫폼 오버라이드
            platform = _safe(r[8])
        cap_kst = (r[1] + timedelta(hours=9)) if r[1] else None
        ws1.append([
            media, cat, cat_detail, "AdScope",
            _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]),
            app_media, media_platform, placement, platform,
            cap_kst.year if cap_kst else "",
            cap_kst.month if cap_kst else "",
            _rad_date(r[9]), _rad_date(r[10]),
            round(r[11] or 0),
        ])
    for row in ws1.iter_rows(min_row=2, min_col=18, max_col=18):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws1, len(h1))
    ws1.auto_filter.ref = ws1.dimensions

    # --- Sheet 2: 쇼핑광고 (18컬럼 + 쇼핑전용 3컬럼 = 21컬럼) ---
    ws_shop = wb.create_sheet("쇼핑광고")
    h_shop = ["매체", "카테고리", "카테고리 세분", "구분",
              "광고주", "브랜드/제품",
              "업종(대)", "업종(중)", "업종(소)",
              "앱/매체", "매체 카테고리/플랫폼", "게재면", "플랫폼",
              "연", "월", "시작일", "종료일", "광고비(전원)",
              "검색키워드", "상품가격(원)", "쇼핑카테고리"]
    ws_shop.append(h_shop)
    _style_header(ws_shop, len(h_shop))

    spend_sub_shop = (
        select(
            Campaign.advertiser_id,
            func.sum(SpendEstimate.est_daily_spend).label("total_spend"),
        )
        .join(SpendEstimate, SpendEstimate.campaign_id == Campaign.id)
        .where(
            Campaign.channel == "naver_shopping",
            SpendEstimate.date >= date_from,
            SpendEstimate.date <= date_to,
        )
        .group_by(Campaign.advertiser_id)
        .subquery()
    )

    q_shop = (
        select(
            AdSnapshot.channel,
            AdSnapshot.captured_at,
            AdDetail.advertiser_name_raw,
            AdDetail.brand,
            Industry.industry_large,
            Industry.industry_medium,
            Industry.name,
            AdDetail.ad_placement,
            AdDetail.first_seen_at,
            AdDetail.last_seen_at,
            spend_sub_shop.c.total_spend,
            AdDetail.extra_data,
        )
        .join(AdSnapshot, AdDetail.snapshot_id == AdSnapshot.id)
        .outerjoin(Advertiser, AdDetail.advertiser_id == Advertiser.id)
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .outerjoin(spend_sub_shop, spend_sub_shop.c.advertiser_id == AdDetail.advertiser_id)
        .where(
            AdSnapshot.channel == "naver_shopping",
            AdSnapshot.captured_at >= date_from,
            AdSnapshot.captured_at <= date_to,
        )
        .order_by(AdSnapshot.captured_at.desc())
    )
    if advertiser_id:
        q_shop = q_shop.where(AdDetail.advertiser_id == advertiser_id)

    for r in (await db.execute(q_shop)).all():
        ch = _safe(r[0])
        media, cat, cat_detail, app_media, media_platform, platform = _ch(ch)
        placement = _PLACEMENT_MAP.get(_safe(r[7]), _safe(r[7]))
        cap_kst = (r[1] + timedelta(hours=9)) if r[1] else None
        extra = r[11] or {}
        keyword = _safe(extra.get("keyword", ""))
        price = extra.get("price", "")
        shopping_cat = _safe(extra.get("shopping_category", extra.get("category", "")))
        ws_shop.append([
            media, cat, cat_detail, "AdScope",
            _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]),
            app_media, media_platform, placement, platform,
            cap_kst.year if cap_kst else "",
            cap_kst.month if cap_kst else "",
            _rad_date(r[8]), _rad_date(r[9]),
            round(r[10] or 0),
            keyword, price, shopping_cat,
        ])
    for col_idx in (18, 20):
        for row in ws_shop.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0'
    _auto_width(ws_shop, len(h_shop))
    ws_shop.auto_filter.ref = ws_shop.dimensions

    # --- Sheet 3: 광고주 ---
    ws2 = wb.create_sheet("광고주")
    h2 = ["ID", "광고주명", "브랜드명", "업종(대)", "업종(중)", "업종(소)", "웹사이트",
          "광고주유형", "등록일"]
    ws2.append(h2)
    _style_header(ws2, len(h2))

    q2 = (
        select(
            Advertiser.id, Advertiser.name, Advertiser.brand_name,
            Industry.industry_large, Industry.industry_medium, Industry.name,
            Advertiser.website, Advertiser.advertiser_type, Advertiser.created_at,
        )
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .order_by(Advertiser.name)
    )
    if advertiser_id:
        q2 = q2.where(Advertiser.id == advertiser_id)

    for r in (await db.execute(q2)).all():
        ws2.append([r[0], _safe(r[1]), _safe(r[2]),
                     _safe(r[3]), _safe(r[4]), _safe(r[5]),
                     _safe(r[6]), _safe(r[7]), _kst_str(r[8])])
    _auto_width(ws2, len(h2))
    ws2.auto_filter.ref = ws2.dimensions

    # --- Sheet 4: 캠페인 ---
    ws3 = wb.create_sheet("캠페인")
    h3 = ["ID", "광고주명", "캠페인명", "채널", "최초발견", "최근발견",
          "활성여부", "총추정광고비", "스냅샷수", "제품/서비스", "목적", "상태"]
    ws3.append(h3)
    _style_header(ws3, len(h3))

    q3 = (
        select(
            Campaign.id, Advertiser.name, Campaign.campaign_name,
            Campaign.channel, Campaign.first_seen, Campaign.last_seen,
            Campaign.is_active, Campaign.total_est_spend, Campaign.snapshot_count,
            Campaign.product_service, Campaign.objective, Campaign.status,
        )
        .join(Advertiser, Campaign.advertiser_id == Advertiser.id)
        .order_by(Campaign.last_seen.desc())
    )
    if advertiser_id:
        q3 = q3.where(Campaign.advertiser_id == advertiser_id)
    if channel:
        q3 = q3.where(Campaign.channel == channel)

    for r in (await db.execute(q3)).all():
        ws3.append([
            r[0], _safe(r[1]), _safe(r[2]), _safe(r[3]),
            _kst_str(r[4]), _kst_str(r[5]),
            "Y" if r[6] else "N", r[7] or 0, r[8] or 0,
            _safe(r[9]), _safe(r[10]), _safe(r[11]),
        ])
    for row in ws3.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws3, len(h3))
    ws3.auto_filter.ref = ws3.dimensions

    # --- Sheet 5: 광고비추정 ---
    ws4 = wb.create_sheet("광고비추정")
    h4 = ["날짜", "채널", "광고주명", "캠페인ID", "일추정광고비", "신뢰도", "산출방법"]
    ws4.append(h4)
    _style_header(ws4, len(h4))

    q4 = (
        select(
            SpendEstimate.date, SpendEstimate.channel,
            Advertiser.name, SpendEstimate.campaign_id,
            SpendEstimate.est_daily_spend, SpendEstimate.confidence,
            SpendEstimate.calculation_method,
        )
        .join(Campaign, SpendEstimate.campaign_id == Campaign.id)
        .join(Advertiser, Campaign.advertiser_id == Advertiser.id)
        .where(SpendEstimate.date >= date_from, SpendEstimate.date <= date_to)
        .order_by(SpendEstimate.date.desc())
    )
    if advertiser_id:
        q4 = q4.where(Campaign.advertiser_id == advertiser_id)
    if channel:
        q4 = q4.where(SpendEstimate.channel == channel)

    for r in (await db.execute(q4)).all():
        ws4.append([
            _kst_date_str(r[0]), _safe(r[1]), _safe(r[2]), r[3],
            r[4] or 0, round(r[5] or 0, 2), _safe(r[6]),
        ])
    for row in ws4.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '#,##0'
    _auto_width(ws4, len(h4))
    ws4.auto_filter.ref = ws4.dimensions

    filename = f"adscope_ad_scope_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 4b. GET /api/export/social-scope.xlsx -- Social Scope raw data (multi-sheet)
# ---------------------------------------------------------------------------
@router.get("/social-scope.xlsx")
async def export_social_scope_xlsx(
    platform: str | None = None,
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Social Scope raw data -- 3 sheets: channels, contents, stats trend."""
    date_from, date_to = _default_date_range(date_from, date_to)
    ck = _cache_key("social-scope", platform, advertiser_id,
                     _kst_date_str(date_from), _kst_date_str(date_to))

    wb = Workbook()

    # --- Sheet 1: 채널목록 (latest per channel) ---
    ws1 = wb.active
    ws1.title = "채널목록"
    h1 = ["광고주ID", "광고주명", "플랫폼", "채널URL", "구독자",
          "팔로워", "총게시물수", "총조회수", "평균좋아요",
          "평균조회수", "인게이지먼트율", "수집일시"]
    ws1.append(h1)
    _style_header(ws1, len(h1))

    latest_sub = (
        select(
            ChannelStats.advertiser_id,
            ChannelStats.platform,
            func.max(ChannelStats.id).label("max_id"),
        )
        .group_by(ChannelStats.advertiser_id, ChannelStats.platform)
        .subquery()
    )
    q1 = (
        select(
            ChannelStats.advertiser_id, Advertiser.name,
            ChannelStats.platform, ChannelStats.channel_url,
            ChannelStats.subscribers, ChannelStats.followers,
            ChannelStats.total_posts, ChannelStats.total_views,
            ChannelStats.avg_likes, ChannelStats.avg_views,
            ChannelStats.engagement_rate, ChannelStats.collected_at,
        )
        .join(latest_sub, ChannelStats.id == latest_sub.c.max_id)
        .join(Advertiser, ChannelStats.advertiser_id == Advertiser.id)
        .order_by(Advertiser.name)
    )
    if platform:
        q1 = q1.where(ChannelStats.platform == platform)
    if advertiser_id:
        q1 = q1.where(ChannelStats.advertiser_id == advertiser_id)

    for r in (await db.execute(q1)).all():
        ws1.append([
            r[0], _safe(r[1]), _safe(r[2]), _safe(r[3]),
            r[4] or 0, r[5] or 0, r[6] or 0, r[7] or 0,
            round(r[8] or 0, 1), round(r[9] or 0, 1),
            round(r[10] or 0, 2), _kst_str(r[11]),
        ])
    for col_idx in (5, 6, 7, 8):
        for row in ws1.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0'
    _auto_width(ws1, len(h1))
    ws1.auto_filter.ref = ws1.dimensions

    # --- Sheet 2: 콘텐츠목록 (소셜 15컬럼 표준) ---
    ws2 = wb.create_sheet("콘텐츠목록")
    h2 = ["플랫폼", "채널명", "광고주", "브랜드",
          "업종(대)", "업종(중)", "업종(소)",
          "콘텐츠 유형", "제목", "업로드일",
          "조회수", "좋아요", "댓글수",
          "구독자/팔로워", "광고콘텐츠 여부", "URL"]
    ws2.append(h2)
    _style_header(ws2, len(h2))

    # 최신 채널 통계 서브쿼리 (구독자/팔로워)
    latest_stats_sub = (
        select(
            ChannelStats.advertiser_id,
            ChannelStats.platform,
            func.max(ChannelStats.id).label("max_id"),
        )
        .group_by(ChannelStats.advertiser_id, ChannelStats.platform)
        .subquery()
    )
    stats_sub = (
        select(
            ChannelStats.advertiser_id,
            ChannelStats.platform,
            ChannelStats.subscribers,
            ChannelStats.followers,
        )
        .join(latest_stats_sub, ChannelStats.id == latest_stats_sub.c.max_id)
        .subquery()
    )

    q2 = (
        select(
            BrandChannelContent.platform,
            BrandChannelContent.channel_url,
            Advertiser.name,
            Advertiser.brand_name,
            Industry.industry_large,
            Industry.industry_medium,
            Industry.name,
            BrandChannelContent.content_type,
            BrandChannelContent.title,
            BrandChannelContent.upload_date,
            BrandChannelContent.view_count,
            BrandChannelContent.like_count,
            BrandChannelContent.comment_count,
            stats_sub.c.subscribers,
            stats_sub.c.followers,
            BrandChannelContent.is_ad_content,
            BrandChannelContent.content_id,
        )
        .join(Advertiser, BrandChannelContent.advertiser_id == Advertiser.id)
        .outerjoin(Industry, Advertiser.industry_id == Industry.id)
        .outerjoin(
            stats_sub,
            (stats_sub.c.advertiser_id == BrandChannelContent.advertiser_id)
            & (stats_sub.c.platform == BrandChannelContent.platform),
        )
        .where(BrandChannelContent.discovered_at >= date_from)
        .where(BrandChannelContent.discovered_at <= date_to)
        .order_by(BrandChannelContent.upload_date.desc())
    )
    if platform:
        q2 = q2.where(BrandChannelContent.platform == platform)
    if advertiser_id:
        q2 = q2.where(BrandChannelContent.advertiser_id == advertiser_id)

    for r in (await db.execute(q2)).all():
        plat = _safe(r[0])
        cid = _safe(r[16])
        if plat == "youtube" and cid:
            url = f"https://www.youtube.com/watch?v={cid}"
        elif plat == "instagram" and cid:
            url = f"https://www.instagram.com/p/{cid}/"
        else:
            url = ""
        # 플랫폼 한국어 라벨
        plat_label = {"youtube": "유튜브", "instagram": "인스타그램"}.get(plat, plat)
        # 콘텐츠 유형 한국어
        ct_map = {"video": "영상", "short": "쇼츠", "reel": "릴스", "post": "포스트"}
        ct_label = ct_map.get(_safe(r[7]), _safe(r[7]))
        # 구독자/팔로워 (유튜브=구독자, 인스타=팔로워)
        sub_fol = r[13] or r[14] or 0
        ws2.append([
            plat_label, _safe(r[1]), _safe(r[2]), _safe(r[3]),
            _safe(r[4]), _safe(r[5]), _safe(r[6]),
            ct_label, _safe(r[8]), _kst_date_str(r[9]),
            r[10] or 0, r[11] or 0, r[12] or 0,
            sub_fol, "Y" if r[15] else "N", url,
        ])
    for col_idx in (11, 12, 13, 14):
        for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0'
    _auto_width(ws2, len(h2))
    ws2.auto_filter.ref = ws2.dimensions

    # --- Sheet 3: 채널통계추이 ---
    ws3 = wb.create_sheet("채널통계추이")
    h3 = ["광고주ID", "광고주명", "플랫폼", "채널URL", "구독자",
          "팔로워", "총게시물수", "총조회수", "인게이지먼트율", "수집일시"]
    ws3.append(h3)
    _style_header(ws3, len(h3))

    q3 = (
        select(
            ChannelStats.advertiser_id, Advertiser.name,
            ChannelStats.platform, ChannelStats.channel_url,
            ChannelStats.subscribers, ChannelStats.followers,
            ChannelStats.total_posts, ChannelStats.total_views,
            ChannelStats.engagement_rate, ChannelStats.collected_at,
        )
        .join(Advertiser, ChannelStats.advertiser_id == Advertiser.id)
        .where(ChannelStats.collected_at >= date_from, ChannelStats.collected_at <= date_to)
        .order_by(ChannelStats.collected_at.desc())
    )
    if platform:
        q3 = q3.where(ChannelStats.platform == platform)
    if advertiser_id:
        q3 = q3.where(ChannelStats.advertiser_id == advertiser_id)

    for r in (await db.execute(q3)).all():
        ws3.append([
            r[0], _safe(r[1]), _safe(r[2]), _safe(r[3]),
            r[4] or 0, r[5] or 0, r[6] or 0, r[7] or 0,
            round(r[8] or 0, 2), _kst_str(r[9]),
        ])
    for col_idx in (5, 6, 7, 8):
        for row in ws3.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0'
    _auto_width(ws3, len(h3))
    ws3.auto_filter.ref = ws3.dimensions

    filename = f"adscope_social_scope_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)


# ---------------------------------------------------------------------------
# 4c. GET /api/export/shop-scope.xlsx -- Shop Scope raw data (multi-sheet)
# ---------------------------------------------------------------------------
@router.get("/shop-scope.xlsx")
async def export_shop_scope_xlsx(
    advertiser_id: int | None = None,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Shop Scope raw data -- 2 sheets: tracked products, snapshots."""
    date_from, date_to = _default_date_range(date_from, date_to)
    ck = _cache_key("shop-scope", advertiser_id,
                     _kst_date_str(date_from), _kst_date_str(date_to))

    wb = Workbook()

    # --- Sheet 1: 추적상품 ---
    ws1 = wb.active
    ws1.title = "추적상품"
    h1 = ["ID", "광고주ID", "광고주명", "스토어명", "상품명",
          "상품URL", "라벨", "활성여부", "등록일"]
    ws1.append(h1)
    _style_header(ws1, len(h1))

    q1 = (
        select(
            SmartStoreTrackedProduct.id, SmartStoreTrackedProduct.advertiser_id,
            Advertiser.name, SmartStoreTrackedProduct.store_name,
            SmartStoreTrackedProduct.product_name, SmartStoreTrackedProduct.product_url,
            SmartStoreTrackedProduct.label, SmartStoreTrackedProduct.is_active,
            SmartStoreTrackedProduct.created_at,
        )
        .outerjoin(Advertiser, SmartStoreTrackedProduct.advertiser_id == Advertiser.id)
        .order_by(SmartStoreTrackedProduct.id)
    )
    if advertiser_id:
        q1 = q1.where(SmartStoreTrackedProduct.advertiser_id == advertiser_id)

    for r in (await db.execute(q1)).all():
        ws1.append([
            r[0], r[1], _safe(r[2]), _safe(r[3]), _safe(r[4]),
            _safe(r[5]), _safe(r[6]), "Y" if r[7] else "N", _kst_str(r[8]),
        ])
    _auto_width(ws1, len(h1))
    ws1.auto_filter.ref = ws1.dimensions

    # --- Sheet 2: 상품스냅샷 ---
    ws2 = wb.create_sheet("상품스냅샷")
    h2 = ["ID", "광고주ID", "광고주명", "스토어명", "상품명", "상품URL",
          "리뷰수", "리뷰증감", "평균평점", "가격", "할인율",
          "랭킹순위", "랭킹카테고리", "위시리스트", "QA수",
          "판매수준", "누적구매수", "구매증감", "추정일판매량",
          "추정방법", "카테고리", "판매자등급", "수집일시"]
    ws2.append(h2)
    _style_header(ws2, len(h2))

    q2 = (
        select(
            SmartStoreSnapshot.id, SmartStoreSnapshot.advertiser_id,
            Advertiser.name, SmartStoreSnapshot.store_name,
            SmartStoreSnapshot.product_name, SmartStoreSnapshot.product_url,
            SmartStoreSnapshot.review_count, SmartStoreSnapshot.review_delta,
            SmartStoreSnapshot.avg_rating, SmartStoreSnapshot.price,
            SmartStoreSnapshot.discount_pct, SmartStoreSnapshot.ranking_position,
            SmartStoreSnapshot.ranking_category, SmartStoreSnapshot.wishlist_count,
            SmartStoreSnapshot.qa_count, SmartStoreSnapshot.estimated_sales_level,
            SmartStoreSnapshot.purchase_cnt, SmartStoreSnapshot.purchase_cnt_delta,
            SmartStoreSnapshot.estimated_daily_sales, SmartStoreSnapshot.estimation_method,
            SmartStoreSnapshot.category_name, SmartStoreSnapshot.seller_grade,
            SmartStoreSnapshot.captured_at,
        )
        .join(Advertiser, SmartStoreSnapshot.advertiser_id == Advertiser.id)
        .where(SmartStoreSnapshot.captured_at >= date_from, SmartStoreSnapshot.captured_at <= date_to)
        .order_by(SmartStoreSnapshot.captured_at.desc())
    )
    if advertiser_id:
        q2 = q2.where(SmartStoreSnapshot.advertiser_id == advertiser_id)

    for r in (await db.execute(q2)).all():
        ws2.append([
            r[0], r[1], _safe(r[2]), _safe(r[3]), _safe(r[4]), _safe(r[5]),
            r[6] or 0, r[7] or 0, round(r[8] or 0, 1), r[9] or 0,
            round(r[10] or 0, 1), r[11] or 0, _safe(r[12]),
            r[13] or 0, r[14] or 0, _safe(r[15]),
            r[16] or 0, r[17] or 0, r[18] or 0,
            _safe(r[19]), _safe(r[20]), _safe(r[21]), _kst_str(r[22]),
        ])
    for col_idx in (7, 10, 14, 17, 19):
        for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0'
    _auto_width(ws2, len(h2))
    ws2.auto_filter.ref = ws2.dimensions

    filename = f"adscope_shop_scope_{_today_str()}.xlsx"
    return _xlsx_response(wb, filename, ck)
