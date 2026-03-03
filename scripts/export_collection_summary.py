"""DB 수집 현황 엑셀 요약 리포트 생성."""
import sqlite3
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "adscope.db")
OUT_PATH = os.path.join(os.path.expanduser("~"), "Desktop", "adscope_collection_summary.xlsx")

conn = sqlite3.connect(DB_PATH)
c = conn.cursor()
wb = Workbook()

# Styles
hdr_font = Font(bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
num_fmt = '#,##0'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
bold_font = Font(bold=True)
total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

def style_header(ws, row=1, col_count=10):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

def style_body(ws, num_cols=None):
    if num_cols is None:
        num_cols = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if cell.column in num_cols:
                cell.number_format = num_fmt

# ═══════════════════════════════════════
# Sheet 1: 매체별 수집 현황 요약
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "1.매체별 수집 요약"
headers1 = ["채널", "수집 소재수", "스냅샷수", "광고주수", "캠페인수",
            "수집 시작일", "수집 종료일", "수집일수",
            "spend_est건수", "30일추정광고비(원)", "캠페인당평균(원)"]
ws1.append(headers1)
style_header(ws1, 1, len(headers1))

c.execute('''
    SELECT s.channel,
           COUNT(DISTINCT d.id) as unique_ads,
           COUNT(*) as total_details,
           COUNT(DISTINCT d.advertiser_id) as unique_advertisers,
           MIN(s.captured_at) as first_cap,
           MAX(s.captured_at) as last_cap
    FROM ad_details d
    JOIN ad_snapshots s ON d.snapshot_id = s.id
    GROUP BY s.channel
    ORDER BY unique_ads DESC
''')
channel_data = c.fetchall()

c.execute('''
    SELECT channel, COUNT(*) as camps,
           ROUND(SUM(total_est_spend), 0),
           ROUND(AVG(CASE WHEN total_est_spend > 0 THEN total_est_spend END), 0)
    FROM campaigns GROUP BY channel
''')
camp_map = {r[0]: (r[1], r[2], r[3]) for r in c.fetchall()}

c.execute('SELECT channel, COUNT(*) FROM spend_estimates GROUP BY channel')
se_map = {r[0]: r[1] for r in c.fetchall()}

total_ads = 0
total_spend = 0
for r in channel_data:
    ch = r[0]
    camps, spend30, avg_spend = camp_map.get(ch, (0, 0, 0))
    se_cnt = se_map.get(ch, 0)
    fc = r[4][:10] if r[4] else ""
    lc = r[5][:10] if r[5] else ""
    try:
        d1 = datetime.strptime(fc, "%Y-%m-%d")
        d2 = datetime.strptime(lc, "%Y-%m-%d")
        days = (d2 - d1).days + 1
    except Exception:
        days = 0
    ws1.append([ch, r[1], r[2], r[3], camps, fc, lc, days, se_cnt, spend30 or 0, avg_spend or 0])
    total_ads += r[1]
    total_spend += (spend30 or 0)

# Total row
row_n = ws1.max_row + 1
for col_idx in range(1, len(headers1) + 1):
    ws1.cell(row=row_n, column=col_idx).fill = total_fill
    ws1.cell(row=row_n, column=col_idx).font = bold_font
    ws1.cell(row=row_n, column=col_idx).border = thin_border
ws1.cell(row=row_n, column=1, value="합계")
ws1.cell(row=row_n, column=2, value=total_ads)
ws1.cell(row=row_n, column=10, value=total_spend)
ws1.cell(row=row_n, column=10).number_format = num_fmt

style_body(ws1, {2, 3, 4, 5, 8, 9, 10, 11})
auto_width(ws1)

# ═══════════════════════════════════════
# Sheet 2: 추정 방법별 상세
# ═══════════════════════════════════════
ws2 = wb.create_sheet("2.추정방법별 상세")
headers2 = ["추정 방법", "채널", "건수", "합계 광고비(원)", "평균 광고비(원)", "비율"]
ws2.append(headers2)
style_header(ws2, 1, len(headers2))

c.execute('''
    SELECT calculation_method, channel, COUNT(*),
           ROUND(SUM(est_daily_spend), 0), ROUND(AVG(est_daily_spend), 0)
    FROM spend_estimates
    GROUP BY calculation_method, channel
    ORDER BY SUM(est_daily_spend) DESC
''')
total_se_spend = 0
rows_data = c.fetchall()
for r in rows_data:
    total_se_spend += r[3]
for r in rows_data:
    ratio = r[3] / total_se_spend if total_se_spend else 0
    ws2.append([r[0], r[1], r[2], r[3], r[4], ratio])

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.border = thin_border
        if cell.column in (3, 4, 5):
            cell.number_format = num_fmt
        if cell.column == 6:
            cell.number_format = pct_fmt
auto_width(ws2)

# ═══════════════════════════════════════
# Sheet 3: ad_hits 분포
# ═══════════════════════════════════════
ws3 = wb.create_sheet("3.ad_hits 분포")
headers3 = ["채널", "1hit", "2hits", "3hits", "4hits", "5+hits", "총건수", "평균hits"]
ws3.append(headers3)
style_header(ws3, 1, len(headers3))

c.execute('SELECT channel, factors FROM spend_estimates WHERE factors IS NOT NULL')
hits_data = {}
for r in c.fetchall():
    ch = r[0]
    try:
        f = json.loads(r[1])
        ah = f.get('ad_hits', 0)
        if ch not in hits_data:
            hits_data[ch] = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 'total': 0, 'sum_hits': 0}
        bucket = min(ah, 5)
        hits_data[ch][bucket] += 1
        hits_data[ch]['total'] += 1
        hits_data[ch]['sum_hits'] += ah
    except Exception:
        pass

for ch in sorted(hits_data.keys()):
    d = hits_data[ch]
    avg_h = d['sum_hits'] / d['total'] if d['total'] else 0
    ws3.append([ch, d.get(1, 0), d.get(2, 0), d.get(3, 0), d.get(4, 0), d.get(5, 0), d['total'], round(avg_h, 1)])
style_body(ws3, {2, 3, 4, 5, 6, 7})
auto_width(ws3)

# ═══════════════════════════════════════
# Sheet 4: 캠페인 상위 100
# ═══════════════════════════════════════
ws4 = wb.create_sheet("4.캠페인 상위100")
headers4 = ["순위", "캠페인ID", "광고주명", "채널", "캠페인명", "30일추정(원)",
            "관측일수", "일평균(원)", "첫수집일", "마지막수집일"]
ws4.append(headers4)
style_header(ws4, 1, len(headers4))

c.execute('''
    SELECT c.id, a.name, c.channel, c.campaign_name, c.total_est_spend,
           CASE WHEN c.first_seen IS NOT NULL AND c.last_seen IS NOT NULL
           THEN CAST(julianday(c.last_seen) - julianday(c.first_seen) + 1 AS INTEGER)
           ELSE 1 END as obs_days,
           c.first_seen, c.last_seen
    FROM campaigns c
    JOIN advertisers a ON c.advertiser_id = a.id
    WHERE c.total_est_spend > 0
    ORDER BY c.total_est_spend DESC
    LIMIT 100
''')
for i, r in enumerate(c.fetchall(), 1):
    daily = round(r[4] / 30) if r[4] else 0
    ws4.append([i, r[0], r[1], r[2], r[3], r[4], r[5] or 1, daily,
                str(r[6])[:10] if r[6] else "", str(r[7])[:10] if r[7] else ""])
style_body(ws4, {6, 7, 8})
auto_width(ws4)

# ═══════════════════════════════════════
# Sheet 5: 산업별 광고주 분포
# ═══════════════════════════════════════
ws5 = wb.create_sheet("5.산업별 광고주")
headers5 = ["산업명", "광고주수", "캠페인수", "30일추정합계(원)", "평균광고비(원)"]
ws5.append(headers5)
style_header(ws5, 1, len(headers5))

c.execute('''
    SELECT COALESCE(i.name, '미분류') as ind_name,
           COUNT(DISTINCT a.id) as adv_cnt,
           COUNT(DISTINCT c.id) as camp_cnt,
           ROUND(SUM(c.total_est_spend), 0) as total_spend,
           ROUND(AVG(CASE WHEN c.total_est_spend > 0 THEN c.total_est_spend END), 0)
    FROM advertisers a
    LEFT JOIN industries i ON a.industry_id = i.id
    LEFT JOIN campaigns c ON c.advertiser_id = a.id
    GROUP BY ind_name
    ORDER BY total_spend DESC
''')
for r in c.fetchall():
    ws5.append([r[0], r[1], r[2], r[3] or 0, r[4] or 0])
style_body(ws5, {2, 3, 4, 5})
auto_width(ws5)

# ═══════════════════════════════════════
# Sheet 6: 현재 추정 파라미터
# ═══════════════════════════════════════
ws6 = wb.create_sheet("6.현재 추정 파라미터")
headers6 = ["채널", "CPC(원)", "Market Calibration", "Inventory Weight",
            "DA감쇠(x0.2)", "hits->clicks(1h)", "hits->clicks(5+h)",
            "현재30일추정(원)", "시장점유율참고"]
ws6.append(headers6)
style_header(ws6, 1, len(headers6))

c.execute('SELECT channel, ROUND(SUM(total_est_spend), 0) FROM campaigns GROUP BY channel')
spend_by_ch = {r[0]: r[1] for r in c.fetchall()}

params = [
    ["naver_search", 500, 0.5, 1.0, "X", 40, 750, spend_by_ch.get("naver_search", 0), "25-30% (2.5~3.0조)"],
    ["naver_da", 800, 0.6, 1.3, "O(x0.2=8/150)", 8, 150, spend_by_ch.get("naver_da", 0), "(네이버 DA 포함)"],
    ["google_gdn", 200, 0.7, 0.6, "O(x0.2=8/150)", 8, 150, spend_by_ch.get("google_gdn", 0), "GDN ~0.4조"],
    ["google_search_ads", 800, 1.0, 0.9, "X(고정50)", 50, 50, spend_by_ch.get("google_search_ads", 0), "구글검색 ~0.7조"],
    ["youtube_ads", "CPV 35", 8.5, 1.5, "조회수기반", "-", "-", spend_by_ch.get("youtube_ads", 0), "유튜브 2.0~2.5조"],
    ["facebook", 700, 3.0, 1.2, "X", 40, 750, spend_by_ch.get("facebook", 0), "메타 2.0~2.4조(FB+IG)"],
    ["instagram", 700, 4.0, 1.3, "X", 40, 750, spend_by_ch.get("instagram", 0), "(메타 내 IG)"],
    ["kakao_da", 600, 1.5, 0.4, "O(x0.2=8/150)", 8, 150, spend_by_ch.get("kakao_da", 0), "카카오 1.0~1.5조"],
    ["naver_shopping", 300, 1.5, 1.1, "X", 40, 750, spend_by_ch.get("naver_shopping", 0), "네이버쇼핑 ~1.1조"],
    ["tiktok_ads", "-", 5.0, 0.8, "X", "-", "-", spend_by_ch.get("tiktok_ads", 0), "틱톡 ~0.3조"],
    ["meta", "-", "-", "-", "접촉기반", 40, 750, spend_by_ch.get("meta", 0), "(fast_crawl 접촉)"],
]
for p in params:
    ws6.append(p)
for row in ws6.iter_rows(min_row=2, max_row=ws6.max_row):
    for cell in row:
        cell.border = thin_border
        if cell.column == 8:
            cell.number_format = num_fmt
auto_width(ws6)

# ═══════════════════════════════════════
# Sheet 7: 일별 수집 추이
# ═══════════════════════════════════════
ws7 = wb.create_sheet("7.일별 수집 추이")
headers7 = ["날짜", "스냅샷수", "소재수", "채널수", "광고주수"]
ws7.append(headers7)
style_header(ws7, 1, len(headers7))

c.execute('''
    SELECT DATE(s.captured_at) as dt,
           COUNT(DISTINCT s.id) as snaps,
           COUNT(DISTINCT d.id) as ads,
           COUNT(DISTINCT s.channel) as channels,
           COUNT(DISTINCT d.advertiser_id) as advs
    FROM ad_snapshots s
    JOIN ad_details d ON d.snapshot_id = s.id
    GROUP BY dt
    ORDER BY dt
''')
for r in c.fetchall():
    ws7.append([r[0], r[1], r[2], r[3], r[4]])
style_body(ws7, {2, 3, 4, 5})
auto_width(ws7)

# ═══════════════════════════════════════
# Sheet 8: serpapi_ads 현황
# ═══════════════════════════════════════
ws8 = wb.create_sheet("8.serpapi_ads")
headers8 = ["포맷", "건수", "광고주수"]
ws8.append(headers8)
style_header(ws8, 1, len(headers8))

c.execute('''
    SELECT format, COUNT(*), COUNT(DISTINCT advertiser_name)
    FROM serpapi_ads
    GROUP BY format
    ORDER BY COUNT(*) DESC
''')
for r in c.fetchall():
    ws8.append([r[0] or "NULL", r[1], r[2]])
style_body(ws8, {2, 3})
auto_width(ws8)

# ═══════════════════════════════════════
# Sheet 9: 채널별 일별 수집
# ═══════════════════════════════════════
ws9 = wb.create_sheet("9.채널x일별 수집")
headers9 = ["날짜", "naver_search", "naver_da", "google_gdn", "google_search_ads",
            "youtube_ads", "kakao_da", "facebook", "instagram", "naver_shopping", "tiktok_ads"]
ws9.append(headers9)
style_header(ws9, 1, len(headers9))

c.execute('''
    SELECT DATE(s.captured_at) as dt, s.channel, COUNT(DISTINCT d.id) as ads
    FROM ad_snapshots s
    JOIN ad_details d ON d.snapshot_id = s.id
    GROUP BY dt, s.channel
    ORDER BY dt
''')
daily_ch = {}
for r in c.fetchall():
    if r[0] not in daily_ch:
        daily_ch[r[0]] = {}
    daily_ch[r[0]][r[1]] = r[2]

ch_order = ["naver_search", "naver_da", "google_gdn", "google_search_ads",
            "youtube_ads", "kakao_da", "facebook", "instagram", "naver_shopping", "tiktok_ads"]
for dt in sorted(daily_ch.keys()):
    row_vals = [dt] + [daily_ch[dt].get(ch, 0) for ch in ch_order]
    ws9.append(row_vals)
style_body(ws9, set(range(2, 12)))
auto_width(ws9)

# Save
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
conn.close()
